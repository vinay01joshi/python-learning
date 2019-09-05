import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("Transaction.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cells = sheet["a:c"]
print(cells)

sheet.append([1005, 5, "$8.76"])
wb.save("UpdatedTransaction.xlsx")

#cell = sheet["a1"]
# cell = sheet.cell(row=1, column=1)
# for row in range(1, sheet.max_row+1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)

# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
#wb.create_sheet("Sheet2", 0)
